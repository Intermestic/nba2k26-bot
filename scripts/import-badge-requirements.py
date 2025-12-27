#!/usr/bin/env python3
"""
Import badge requirements from Excel files into the database
"""
import openpyxl
import mysql.connector
import os
import json
from collections import defaultdict

# Database connection from environment
DB_HOST = os.environ.get('DATABASE_HOST', 'localhost')
DB_USER = os.environ.get('DATABASE_USER', 'root')
DB_PASSWORD = os.environ.get('DATABASE_PASSWORD', '')
DB_NAME = os.environ.get('DATABASE_NAME', 'nba2k26')

def get_db_connection():
    """Get database connection from environment variables"""
    # Parse connection string if available
    db_url = os.environ.get('DATABASE_URL', '')
    if db_url:
        # Format: mysql://user:pass@host:port/dbname?params
        import re
        match = re.match(r'mysql://([^:]+):([^@]+)@([^:]+):(\d+)/([^?]+)', db_url)
        if match:
            user, password, host, port, database = match.groups()
            return mysql.connector.connect(
                host=host,
                port=int(port),
                user=user,
                password=password,
                database=database,
                ssl_disabled=False
            )
    
    return mysql.connector.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME
    )

def read_badge_glossary(filepath):
    """Read badge abbreviations from Badge Glossary sheet"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Badge Glossary']
    
    abbreviations = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # Skip header
            continue
        if not row[0]:  # Skip empty rows
            continue
            
        badge_name = row[0]
        description = row[1] if len(row) > 1 else None
        abbreviation = row[2] if len(row) > 2 else None
        
        if badge_name and abbreviation:
            abbreviations.append({
                'abbreviation': abbreviation.upper(),
                'fullName': badge_name.title(),
                'category': None  # Will be filled from badge caps
            })
    
    return abbreviations

def read_badge_requirements(filepath, sheet_name):
    """Read badge requirements from Badge Caps or Challenger Badge sheet"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]
    
    # Group requirements by badge + tier
    badge_data = defaultdict(lambda: {'attributes': []})
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # Skip header
            continue
        if not row[2]:  # Skip if no badge name
            continue
            
        category = row[1]
        badge_name = row[2]
        attribute = row[3]
        bronze = row[4]
        silver = row[5]
        gold = row[6]
        min_height = row[7] if len(row) > 7 else None
        max_height = row[8] if len(row) > 8 else None
        
        # Store for each tier
        for tier, threshold in [('bronze', bronze), ('silver', silver), ('gold', gold)]:
            if threshold:
                key = f"{badge_name}_{tier}"
                badge_data[key]['badge_name'] = badge_name
                badge_data[key]['tier'] = tier
                badge_data[key]['category'] = category
                badge_data[key]['attributes'].append({
                    'name': attribute,
                    'threshold': threshold
                })
    
    # Convert to list of requirements
    requirements = []
    for key, data in badge_data.items():
        req = {
            'badgeName': data['badge_name'],
            'tier': data['tier'],
            'category': data['category'],
            'attributes': data['attributes']
        }
        requirements.append(req)
    
    return requirements

def insert_abbreviations(conn, abbreviations):
    """Insert badge abbreviations into database"""
    cursor = conn.cursor()
    
    # Clear existing data
    cursor.execute("DELETE FROM badge_abbreviations")
    
    # Insert new data
    for abbr in abbreviations:
        cursor.execute("""
            INSERT INTO badge_abbreviations (abbreviation, fullName, category)
            VALUES (%s, %s, %s)
        """, (abbr['abbreviation'], abbr['fullName'], abbr['category']))
    
    conn.commit()
    print(f"✅ Inserted {len(abbreviations)} badge abbreviations")

def insert_requirements(conn, requirements):
    """Insert badge requirements into database"""
    cursor = conn.cursor()
    
    # Clear existing data
    cursor.execute("DELETE FROM badge_requirements")
    
    # Insert new data
    for req in requirements:
        # Get up to 3 attributes
        attr1 = req['attributes'][0] if len(req['attributes']) > 0 else None
        attr2 = req['attributes'][1] if len(req['attributes']) > 1 else None
        attr3 = req['attributes'][2] if len(req['attributes']) > 2 else None
        
        cursor.execute("""
            INSERT INTO badge_requirements 
            (badgeName, tier, attribute1, threshold1, attribute2, threshold2, attribute3, threshold3)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            req['badgeName'],
            req['tier'],
            attr1['name'] if attr1 else None,
            attr1['threshold'] if attr1 else None,
            attr2['name'] if attr2 else None,
            attr2['threshold'] if attr2 else None,
            attr3['name'] if attr3 else None,
            attr3['threshold'] if attr3 else None,
        ))
    
    conn.commit()
    print(f"✅ Inserted {len(requirements)} badge requirements")

def main():
    print("🔄 Importing badge requirements from Excel files...")
    
    # Read data from Excel files
    print("\n📖 Reading Badge Glossary...")
    abbreviations = read_badge_glossary('/home/ubuntu/upload/HoF_Upgrades_Master_WithGlossary_Final.xlsx')
    
    print("📖 Reading Badge Caps...")
    requirements = read_badge_requirements('/home/ubuntu/upload/HoF_Upgrades_Master_WithGlossary_Final.xlsx', 'Badge Caps')
    
    print("📖 Reading Challenger Requirements...")
    challenger_reqs = read_badge_requirements('/home/ubuntu/upload/Challenger_Requirements.xlsx', 'Challenger Badge')
    requirements.extend(challenger_reqs)
    
    print(f"\n📊 Found:")
    print(f"  - {len(abbreviations)} badge abbreviations")
    print(f"  - {len(requirements)} badge requirements")
    
    # Connect to database
    print("\n🔌 Connecting to database...")
    conn = get_db_connection()
    
    # Insert data
    print("\n💾 Inserting data into database...")
    insert_abbreviations(conn, abbreviations)
    insert_requirements(conn, requirements)
    
    conn.close()
    print("\n✅ Import complete!")

if __name__ == '__main__':
    main()
