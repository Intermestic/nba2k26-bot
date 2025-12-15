import openpyxl
import mysql.connector
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv('/home/ubuntu/nba2k26-database/.env')

# Parse DATABASE_URL
db_url = os.getenv('DATABASE_URL')
# Format: mysql://user:pass@host:port/dbname
parts = db_url.replace('mysql://', '').split('@')
user_pass = parts[0].split(':')
host_port_db = parts[1].split('/')
host_port = host_port_db[0].split(':')

db_config = {
    'user': user_pass[0],
    'password': user_pass[1],
    'host': host_port[0],
    'port': int(host_port[1]) if len(host_port) > 1 else 3306,
    'database': host_port_db[1].split('?')[0]
}

print(f"Connecting to database: {db_config['host']}:{db_config['port']}/{db_config['database']}")

# Connect to database
conn = mysql.connector.connect(**db_config)
cursor = conn.cursor()

# Load Excel file
wb = openpyxl.load_workbook('/home/ubuntu/upload/HoF_Upgrades_Master_WithGlossary_Final(1).xlsx')

# Step 1: Clear existing data
print("\nClearing existing badge data...")
cursor.execute("DELETE FROM badge_requirements")
cursor.execute("DELETE FROM badge_abbreviations")
conn.commit()
print("✓ Cleared badge_requirements and badge_abbreviations tables")

# Step 2: Import Badge Glossary (abbreviations)
print("\nImporting badge abbreviations from Badge Glossary...")
ws_glossary = wb['Badge Glossary']
abbreviations = []
for i, row in enumerate(ws_glossary.iter_rows(values_only=True), 1):
    if i == 1:  # Skip header
        continue
    if row[0] and row[2]:  # Badge name and abbreviation exist
        badge_name = row[0].strip()
        abbreviation = row[2].strip()
        description = row[1].strip() if row[1] else ''
        abbreviations.append((abbreviation, badge_name, description))

cursor.executemany(
    "INSERT INTO badge_abbreviations (abbreviation, fullName, category) VALUES (%s, %s, %s)",
    abbreviations
)
conn.commit()
print(f"✓ Imported {len(abbreviations)} badge abbreviations")

# Step 3: Import Badge Caps (requirements)
print("\nImporting badge requirements from Badge Caps...")
ws_caps = wb['Badge Caps']
requirements = []
for i, row in enumerate(ws_caps.iter_rows(values_only=True), 1):
    if i == 1:  # Skip header
        continue
    if row[2] and row[3]:  # Badge name and attribute requirement exist
        badge_name = row[2].strip()
        attribute = row[3].strip()
        bronze = row[4] if row[4] else None
        silver = row[5] if row[5] else None
        gold = row[6] if row[6] else None
        
        # Insert bronze requirement
        if bronze:
            requirements.append((badge_name, 'bronze', attribute, bronze, None, None))
        
        # Insert silver requirement
        if silver:
            requirements.append((badge_name, 'silver', attribute, silver, None, None))
        
        # Insert gold requirement
        if gold:
            requirements.append((badge_name, 'gold', attribute, gold, None, None))

cursor.executemany(
    "INSERT INTO badge_requirements (badgeName, tier, attribute1, threshold1, attribute2, threshold2) VALUES (%s, %s, %s, %s, %s, %s)",
    requirements
)
conn.commit()
print(f"✓ Imported {len(requirements)} badge requirements")

# Step 4: Verify data
cursor.execute("SELECT COUNT(*) FROM badge_abbreviations")
abbr_count = cursor.fetchone()[0]
cursor.execute("SELECT COUNT(*) FROM badge_requirements")
req_count = cursor.fetchone()[0]

print("\n" + "="*80)
print("IMPORT SUMMARY:")
print(f"  Badge Abbreviations: {abbr_count}")
print(f"  Badge Requirements: {req_count}")
print("="*80)

# Show some examples
print("\nSample abbreviations:")
cursor.execute("SELECT * FROM badge_abbreviations WHERE abbreviation IN ('SS', 'SSS', 'PTZ', 'CHL', 'LR') ORDER BY abbreviation")
for row in cursor.fetchall():
    print(f"  {row[1]} = {row[2]}")

print("\nSample requirements for Shifty Shooter:")
cursor.execute("SELECT * FROM badge_requirements WHERE badgeName = 'Shifty Shooter' ORDER BY tier")
for row in cursor.fetchall():
    print(f"  {row[2]} {row[1]}: {row[3]} {row[4]} (threshold: {row[5]})")

cursor.close()
conn.close()
print("\n✓ Badge import complete!")
